#!/usr/bin/env python3
"""
SDTM Mapping Specification Template Generator
Generates Excel templates for domain mapping specifications.

Usage:
    python generate_mapping_template.py --domain DM --output dm_spec.xlsx
    python generate_mapping_template.py --domain AE --output ae_spec.xlsx
"""

import argparse
import json
import os
import sys

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# SDTM IG 3.4 Variable Definitions
DOMAIN_VARIABLES = {
    'DM': [
        {'variable': 'STUDYID', 'label': 'Study Identifier', 'type': 'Char', 'length': 20, 'core': 'Req', 'origin': 'Assigned'},
        {'variable': 'DOMAIN', 'label': 'Domain Abbreviation', 'type': 'Char', 'length': 2, 'core': 'Req', 'origin': 'Assigned'},
        {'variable': 'USUBJID', 'label': 'Unique Subject Identifier', 'type': 'Char', 'length': 40, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'SUBJID', 'label': 'Subject Identifier for the Study', 'type': 'Char', 'length': 20, 'core': 'Req', 'origin': 'CRF'},
        {'variable': 'RFSTDTC', 'label': 'Subject Reference Start Date/Time', 'type': 'Char', 'length': 19, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'RFENDTC', 'label': 'Subject Reference End Date/Time', 'type': 'Char', 'length': 19, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'RFXSTDTC', 'label': 'Date/Time of First Study Treatment', 'type': 'Char', 'length': 19, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'RFXENDTC', 'label': 'Date/Time of Last Study Treatment', 'type': 'Char', 'length': 19, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'RFICDTC', 'label': 'Date/Time of Informed Consent', 'type': 'Char', 'length': 19, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'RFPENDTC', 'label': 'Date/Time of End of Participation', 'type': 'Char', 'length': 19, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'DTHDTC', 'label': 'Date/Time of Death', 'type': 'Char', 'length': 19, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'DTHFL', 'label': 'Subject Death Flag', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'SITEID', 'label': 'Study Site Identifier', 'type': 'Char', 'length': 20, 'core': 'Req', 'origin': 'CRF'},
        {'variable': 'INVID', 'label': 'Investigator Identifier', 'type': 'Char', 'length': 20, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'INVNAM', 'label': 'Investigator Name', 'type': 'Char', 'length': 200, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'BRTHDTC', 'label': 'Date/Time of Birth', 'type': 'Char', 'length': 19, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'AGE', 'label': 'Age', 'type': 'Num', 'length': 8, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'AGEU', 'label': 'Age Units', 'type': 'Char', 'length': 10, 'core': 'Exp', 'origin': 'Derived', 'codelist': 'AGEU'},
        {'variable': 'SEX', 'label': 'Sex', 'type': 'Char', 'length': 2, 'core': 'Req', 'origin': 'CRF', 'codelist': 'SEX'},
        {'variable': 'RACE', 'label': 'Race', 'type': 'Char', 'length': 60, 'core': 'Exp', 'origin': 'CRF', 'codelist': 'RACE'},
        {'variable': 'ETHNIC', 'label': 'Ethnicity', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'ETHNIC'},
        {'variable': 'ARMCD', 'label': 'Planned Arm Code', 'type': 'Char', 'length': 20, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'ARM', 'label': 'Description of Planned Arm', 'type': 'Char', 'length': 200, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'ACTARMCD', 'label': 'Actual Arm Code', 'type': 'Char', 'length': 20, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'ACTARM', 'label': 'Description of Actual Arm', 'type': 'Char', 'length': 200, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'COUNTRY', 'label': 'Country', 'type': 'Char', 'length': 3, 'core': 'Req', 'origin': 'CRF', 'codelist': 'ISO 3166'},
        {'variable': 'DMDTC', 'label': 'Date/Time of Collection', 'type': 'Char', 'length': 19, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'DMDY', 'label': 'Study Day of Collection', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
    ],
    'AE': [
        {'variable': 'STUDYID', 'label': 'Study Identifier', 'type': 'Char', 'length': 20, 'core': 'Req', 'origin': 'Assigned'},
        {'variable': 'DOMAIN', 'label': 'Domain Abbreviation', 'type': 'Char', 'length': 2, 'core': 'Req', 'origin': 'Assigned'},
        {'variable': 'USUBJID', 'label': 'Unique Subject Identifier', 'type': 'Char', 'length': 40, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'AESEQ', 'label': 'Sequence Number', 'type': 'Num', 'length': 8, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'AESPID', 'label': 'Sponsor-Defined Identifier', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'AETERM', 'label': 'Reported Term for the Adverse Event', 'type': 'Char', 'length': 200, 'core': 'Req', 'origin': 'CRF'},
        {'variable': 'AELLT', 'label': 'Lowest Level Term', 'type': 'Char', 'length': 200, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AELLTCD', 'label': 'Lowest Level Term Code', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEDECOD', 'label': 'Dictionary-Derived Term', 'type': 'Char', 'length': 200, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'AEPTCD', 'label': 'Preferred Term Code', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEHLT', 'label': 'High Level Term', 'type': 'Char', 'length': 200, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEHLTCD', 'label': 'High Level Term Code', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEHLGT', 'label': 'High Level Group Term', 'type': 'Char', 'length': 200, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEHLGTCD', 'label': 'High Level Group Term Code', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEBODSYS', 'label': 'Body System or Organ Class', 'type': 'Char', 'length': 200, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'AESOCCD', 'label': 'Primary SOC Code', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AESEV', 'label': 'Severity/Intensity', 'type': 'Char', 'length': 20, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'AESEV'},
        {'variable': 'AESER', 'label': 'Serious Event', 'type': 'Char', 'length': 1, 'core': 'Exp', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AEACN', 'label': 'Action Taken with Study Treatment', 'type': 'Char', 'length': 50, 'core': 'Exp', 'origin': 'CRF', 'codelist': 'ACN'},
        {'variable': 'AEREL', 'label': 'Causality', 'type': 'Char', 'length': 50, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'AEOUT', 'label': 'Outcome of Adverse Event', 'type': 'Char', 'length': 50, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'OUT'},
        {'variable': 'AESCAN', 'label': 'Involves Cancer', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AESCONG', 'label': 'Congenital Anomaly or Birth Defect', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AESDISAB', 'label': 'Persist or Significant Disability', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AESDTH', 'label': 'Results in Death', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AESHOSP', 'label': 'Requires or Prolongs Hospitalization', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AESLIFE', 'label': 'Is Life Threatening', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AESMIE', 'label': 'Other Medically Important SAE', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AECONTRT', 'label': 'Concomitant Treatment Given', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'AETOXGR', 'label': 'Standard Toxicity Grade', 'type': 'Char', 'length': 5, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'AESTDTC', 'label': 'Start Date/Time of Adverse Event', 'type': 'Char', 'length': 19, 'core': 'Exp', 'origin': 'CRF'},
        {'variable': 'AEENDTC', 'label': 'End Date/Time of Adverse Event', 'type': 'Char', 'length': 19, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'AESTDY', 'label': 'Study Day of Start of AE', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'AEENDY', 'label': 'Study Day of End of AE', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'EPOCH', 'label': 'Epoch', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'Derived', 'codelist': 'EPOCH'},
    ],
    'LB': [
        {'variable': 'STUDYID', 'label': 'Study Identifier', 'type': 'Char', 'length': 20, 'core': 'Req', 'origin': 'Assigned'},
        {'variable': 'DOMAIN', 'label': 'Domain Abbreviation', 'type': 'Char', 'length': 2, 'core': 'Req', 'origin': 'Assigned'},
        {'variable': 'USUBJID', 'label': 'Unique Subject Identifier', 'type': 'Char', 'length': 40, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'LBSEQ', 'label': 'Sequence Number', 'type': 'Num', 'length': 8, 'core': 'Req', 'origin': 'Derived'},
        {'variable': 'LBREFID', 'label': 'Specimen ID', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'LBTESTCD', 'label': 'Lab Test Short Name', 'type': 'Char', 'length': 8, 'core': 'Req', 'origin': 'Assigned', 'codelist': 'LBTESTCD'},
        {'variable': 'LBTEST', 'label': 'Lab Test Name', 'type': 'Char', 'length': 40, 'core': 'Req', 'origin': 'Assigned', 'codelist': 'LBTEST'},
        {'variable': 'LBCAT', 'label': 'Category for Lab Test', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'Assigned'},
        {'variable': 'LBSCAT', 'label': 'Subcategory for Lab Test', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'Assigned'},
        {'variable': 'LBORRES', 'label': 'Result or Finding in Original Units', 'type': 'Char', 'length': 200, 'core': 'Exp', 'origin': 'CRF'},
        {'variable': 'LBORRESU', 'label': 'Original Units', 'type': 'Char', 'length': 40, 'core': 'Exp', 'origin': 'CRF', 'codelist': 'UNIT'},
        {'variable': 'LBORNRLO', 'label': 'Reference Range Lower Limit (Orig)', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'LBORNRHI', 'label': 'Reference Range Upper Limit (Orig)', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'LBSTRESC', 'label': 'Character Result/Finding in Std Format', 'type': 'Char', 'length': 200, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'LBSTRESN', 'label': 'Numeric Result/Finding in Std Units', 'type': 'Num', 'length': 8, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'LBSTRESU', 'label': 'Standard Units', 'type': 'Char', 'length': 40, 'core': 'Exp', 'origin': 'Derived', 'codelist': 'UNIT'},
        {'variable': 'LBSTNRLO', 'label': 'Reference Range Lower Limit (Std)', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'LBSTNRHI', 'label': 'Reference Range Upper Limit (Std)', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'LBNRIND', 'label': 'Reference Range Indicator', 'type': 'Char', 'length': 20, 'core': 'Perm', 'origin': 'Derived', 'codelist': 'NRIND'},
        {'variable': 'LBSPEC', 'label': 'Specimen Type', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'SPECTYPE'},
        {'variable': 'LBMETHOD', 'label': 'Method of Test or Examination', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'CRF'},
        {'variable': 'LBBLFL', 'label': 'Baseline Flag', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'Derived', 'codelist': 'NY'},
        {'variable': 'LBFAST', 'label': 'Fasting Status', 'type': 'Char', 'length': 1, 'core': 'Perm', 'origin': 'CRF', 'codelist': 'NY'},
        {'variable': 'VISITNUM', 'label': 'Visit Number', 'type': 'Num', 'length': 8, 'core': 'Exp', 'origin': 'Derived'},
        {'variable': 'VISIT', 'label': 'Visit Name', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'Derived'},
        {'variable': 'EPOCH', 'label': 'Epoch', 'type': 'Char', 'length': 40, 'core': 'Perm', 'origin': 'Derived', 'codelist': 'EPOCH'},
        {'variable': 'LBDTC', 'label': 'Date/Time of Specimen Collection', 'type': 'Char', 'length': 19, 'core': 'Exp', 'origin': 'CRF'},
        {'variable': 'LBDY', 'label': 'Study Day of Specimen Collection', 'type': 'Num', 'length': 8, 'core': 'Perm', 'origin': 'Derived'},
    ],
}


def create_mapping_template(domain: str, output_path: str) -> bool:
    """Create Excel mapping specification template"""

    if not HAS_OPENPYXL:
        print("Error: openpyxl not installed. Install with: pip install openpyxl")
        return False

    if domain not in DOMAIN_VARIABLES:
        print(f"Error: Domain '{domain}' not found. Available: {', '.join(DOMAIN_VARIABLES.keys())}")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = f"{domain} Mapping"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    req_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    exp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ['Variable', 'Label', 'Type', 'Length', 'Core', 'Origin',
               'Source Dataset', 'Source Variable', 'Derivation Rule', 'Codelist', 'Comments']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    variables = DOMAIN_VARIABLES[domain]
    for row, var in enumerate(variables, 2):
        ws.cell(row=row, column=1, value=var['variable']).border = thin_border
        ws.cell(row=row, column=2, value=var['label']).border = thin_border
        ws.cell(row=row, column=3, value=var['type']).border = thin_border
        ws.cell(row=row, column=4, value=var['length']).border = thin_border
        ws.cell(row=row, column=5, value=var['core']).border = thin_border
        ws.cell(row=row, column=6, value=var['origin']).border = thin_border
        ws.cell(row=row, column=7, value='').border = thin_border  # Source Dataset
        ws.cell(row=row, column=8, value='').border = thin_border  # Source Variable
        ws.cell(row=row, column=9, value='').border = thin_border  # Derivation
        ws.cell(row=row, column=10, value=var.get('codelist', '')).border = thin_border
        ws.cell(row=row, column=11, value='').border = thin_border  # Comments

        # Color coding by Core status
        if var['core'] == 'Req':
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = req_fill
        elif var['core'] == 'Exp':
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = exp_fill

    # Adjust column widths
    widths = [12, 45, 6, 8, 6, 10, 15, 15, 50, 15, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Add legend sheet
    legend = wb.create_sheet("Legend")
    legend.cell(row=1, column=1, value="Color Legend").font = Font(bold=True)
    legend.cell(row=2, column=1, value="Required Variables").fill = req_fill
    legend.cell(row=3, column=1, value="Expected Variables").fill = exp_fill
    legend.cell(row=4, column=1, value="Permissible Variables")

    legend.cell(row=6, column=1, value="Origin Values").font = Font(bold=True)
    legend.cell(row=7, column=1, value="CRF - Collected on Case Report Form")
    legend.cell(row=8, column=1, value="Derived - Calculated or transformed")
    legend.cell(row=9, column=1, value="Assigned - Sponsor-assigned constant")
    legend.cell(row=10, column=1, value="Protocol - From protocol document")
    legend.cell(row=11, column=1, value="eDT - Electronic data transfer")

    # Save workbook
    wb.save(output_path)
    return True


def main():
    parser = argparse.ArgumentParser(description='Generate SDTM mapping specification template')
    parser.add_argument('--domain', required=True, help='SDTM domain (DM, AE, LB, etc.)')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    parser.add_argument('--format', choices=['xlsx', 'json'], default='xlsx',
                        help='Output format (default: xlsx)')

    args = parser.parse_args()

    domain = args.domain.upper()

    if args.format == 'json':
        # Output as JSON
        if domain not in DOMAIN_VARIABLES:
            print(f"Error: Domain '{domain}' not found")
            sys.exit(1)
        with open(args.output, 'w') as f:
            json.dump({'domain': domain, 'variables': DOMAIN_VARIABLES[domain]}, f, indent=2)
        print(f"JSON template created: {args.output}")
    else:
        # Output as Excel
        if create_mapping_template(domain, args.output):
            print(f"Excel template created: {args.output}")
            print(f"  Domain: {domain}")
            print(f"  Variables: {len(DOMAIN_VARIABLES.get(domain, []))}")
        else:
            sys.exit(1)


if __name__ == '__main__':
    main()
